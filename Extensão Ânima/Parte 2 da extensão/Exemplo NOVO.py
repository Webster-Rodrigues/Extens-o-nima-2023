
import pandas as pd
import numpy as np
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Color, Alignment

combustiveis_df = pd.read_excel("BaseDados_principal.xlsx")
combustiveis_df.head()
#print(combustiveis_df)

combustiveis_df["Ativo"] = True
#print(combustiveis_df)

combustiveis_df["Obs"] = ["MELHOR CIDADE" if Municipio == 'SAO PAULO' else None for Municipio in combustiveis_df["Municipio"]]
#print(combustiveis_df.loc[combustiveis_df['Municipio'].isin(['SAO PAULO','INDAIATUBA', 'CAMPINAS', 'SALTO']),
#['Municipio', 'Obs']])


'''Preencher uma coluna 'Valor de Venda - Status' que verifica se o valor de venda for maior que 6,5$ 
ele fala que tá Caro..caso contrário, está barato'''

combustiveis_df["Status do Valor de Venda"] = np.where(combustiveis_df["Valor de Venda"] > 6.5, "Caro", "Barato")
#print(combustiveis_df[["Revenda", "Valor de Venda", "Status do Valor de Venda"]])


num_habitantes_df = pd.read_csv("ibge_num_habitantes_estimado.csv", sep = ';')
#print(num_habitantes_df)

# Calcular postos de gasolina por habitante temos na amostragem de combustiveis nov/2021
num_habitantes_df = pd.read_csv("ibge_num_habitantes_estimado.csv", sep=";")
num_habitantes_df.rename(columns={"Estado":"Estado - Sigla"}, inplace = True)
#print(num_habitantes_df)


# Faz um MERGE dos dois dataframes
colunas = ['Municipio', 'Estado - Sigla']
merge_df = combustiveis_df.merge(num_habitantes_df, how = "inner", on = colunas)
#print(merge_df.info())


#Destruir coluna completamente vazia (todas as linhas são nulas)
merge_df.dropna(axis = 'columns', inplace = True)
#print(merge_df.info())

colunas = ['Regiao - Sigla', 'Nome da Rua', 'Numero Rua',
            'Bairro', 'Cep', 'Produto', 'Data da Coleta', 'Valor de Venda',
            'Unidade de Medida', 'Bandeira', 'Ativo', 'Status do Valor de Venda']
merge_df.drop(labels = colunas, axis = 1, inplace = True)
#print(merge_df.info())


# Remover a linhas duplicadas
merge_df.drop_duplicates(inplace = True)
#print(merge_df.head(100))

#Agrupar e contar quantos postos tem na cidade..
postos_por_municipio_df = merge_df.groupby(by = ['Estado - Sigla', 'Municipio', 'NumHabitantes2021']).count()
postos_por_municipio_df.reset_index(inplace = True)

postos_por_municipio_df.drop('CNPJ da Revenda', axis = 1, inplace = True)
postos_por_municipio_df.rename(columns = {"Revenda": "NumPostos"}, inplace = True)

postos_por_municipio_df['NumHabitantesPorPosto'] = postos_por_municipio_df['NumHabitantes2021'] / postos_por_municipio_df['NumPostos']

#print(postos_por_municipio_df)


#Área para teste de  GRÁFICOS

import matplotlib.pyplot as plt

#Agrupar dados
plt.hist(combustiveis_df["Valor de Venda"])
plt.title("Valor médio de combustíveis - Nov/2021")
plt.xlabel("Preço em REAIS")
plt.ylabel("Quantidade de Coletas")
#Traça a linha vermelha tracejada com o preço médio
plt.axvline(combustiveis_df['Valor de Venda'].mean(), color = 'red', linestyle = 'dashed', linewidth = 5)



#Visualizar o consumo médio
consumo_medio = combustiveis_df["Valor de Venda"].groupby(by = combustiveis_df["Produto"]).mean().round(2)
#print(consumo_medio)


#Área do Gráfico em polegadas
plt.figure(figsize=(7,5))

#Plotar o gráfico
consumo_medio.plot(
    kind="barh",
    xlabel="Tipo de Combustível",
    ylabel="Preço reais/litro",
    title="Média de preços por combustível",
    color="red",
    alpha=0.3
)

#Grade
plt.grid()
#plt.show()

excel = "por_litro.xlsx"
consumo_medio.to_excel(excel, "Sumário")

#Personalizar planilha do excel
wb = load_workbook(excel)
ws = wb['Sumário']
cinzinha = PatternFill("solid", fgColor="CCCCCC")
coords = ['A1', 'B1']
for coord in coords:
  ws[coord].fill = cinzinha

#Onde o preço do combustível for maior ou igual a 6,5 reais (6.5) pinta a fonte de vermelho e deixa negrito...
MAX_ROW = ws.max_row
num_linha = 2
while (num_linha <= MAX_ROW):
  coord = 'B'+str(num_linha) #coord="B{0}".format(num_linha)
  if ws[coord].value >= 6.5:
    ws[coord].font = Font(bold=True, color="FF0000")
  num_linha = num_linha + 1

wb.save(excel)





